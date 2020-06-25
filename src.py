import openpyxl as win32
from re import findall


#print("Введите имя исходного файла:")
#file_input = str(input())
#file_path = "D:\\Заявки\\" + file_input + ".xlsx"

#print("Введите имя конечного файла:")
#file_output = str(input())

file_path = "D:\\Мувик\\Бланк ЛИМА.xlsx"
file_output = "Лима_Тест"

#Columns
BUST = 0
COLOR = 2
CUP = 3
SIZE = [4, 5, 6, 7, 8, 9, 10]

#Variables
row_new = 2
row_num = 5

#Filter
bust_pattern = r'^\w+\s\d+'

buffer = {
    "Бюст": 0,
    "Цвет": 0,
    "Чашка": 0,
    "Размер": 0,
    "Кол-во": 0
}

header = {
    4: "70",
    5: "75",
    6: "80",
    7: "85",
    8: "90",
    9: "95",
    10: "100"
}

symbol = {
    4: "E",
    5: "F",
    6: "G",
    7: "H",
    8: "I",
    9: "J",
    10: "K"
}

#Open input file
wb = win32.load_workbook(filename=file_path)
sheet = wb.worksheets[0]

#Create outpute file
wb_new = win32.Workbook()
sheet_new = wb_new.active

#Create headers
j = 1
for head in buffer.keys():
    sheet_new.cell(1, j).value = head
    j += 1

sheet_new.column_dimensions['A'].width = 15
sheet_new.column_dimensions['B'].width = 10

for row in sheet.iter_rows(min_row=row_num, max_col=11, max_row=19, values_only=False):
    #Save bust number
    if row[BUST] != None:
        result = findall(bust_pattern, str(row[BUST].value))
        if result:
            bust_id = result[0]

    #Save bust color
    if row[COLOR].value != None:
        bust_color = row[COLOR].value

    #Search in row
    for i in SIZE:
        if row[i].value != None and row[i].fill.bgColor != row[3].fill.bgColor:
            #Fill in buffer
            buffer["Бюст"] = bust_id
            buffer["Цвет"] = bust_color
            buffer["Чашка"] = row[CUP].value
            buffer["Размер"] = header[i]
            buffer["Кол-во"] = "Лист1!" + symbol[i] + str(row_num)

            #Translate value to output file
            col_new = 1
            for value in buffer.values():
                sheet_new.cell(row_new, col_new).value = str(value)
                col_new += 1

            row_new += 1

    row_num += 1


#Save output file
try:
    wb_new.save("D:\\Заказы\\" + file_output + ".xlsx")
except PermissionError:
    print("Ошибка сохранения")

wb.close()
wb_new.close()